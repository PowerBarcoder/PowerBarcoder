import csv
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# CSV_FILE_PATH_PREFIX = '/PowerBarcoder/data/result/batchRun/'
CSV_FILE_PATH_PREFIX = 'C:/Users/kwz50/IdeaProjects/PowerBarcoder/data/result/batchRun/'
CSV_FILE_PATH_SUFFIX = '_result/qcResult/qcReport.csv'
# CONCAT_CSV_FILE_PATH = '/PowerBarcoder/data/result/batchRun/matrix_data.csv'
CONCAT_CSV_FILE_PATH = 'C:/Users/kwz50/IdeaProjects/PowerBarcoder/data/result/batchRun/matrix_data.csv'
CONCAT_EXCEL_FILE_PATH = 'C:/Users/kwz50/IdeaProjects/PowerBarcoder/data/result/batchRun/matrix_data.xlsx'

FIRST_ASV_ROW_INDEX = 12 - 1
DADA2_IDENTICAL_ROW_INDEX = 334 - 1
DADA2_IDENTICAL_COL_INDEX = 24 - 1
ASV_BARCODE_COL_INDEX = 1 - 1
ASV_NAME_COL_INDEX = 2 - 1
DADA2_DENOISE_R1_HASH_VALUE_COL_INDEX = 28 - 1
DADA2_DENOISE_R2_HASH_VALUE_COL_INDEX = 32 - 1
DADA2_MERGE_PAIRS_HASH_VALUE_COL_INDEX = 36 - 1
DADA2_10N_CONCAT_HASH_VALUE_COL_INDEX = 40 - 1
HASH_VALUE_COL_NAME_LIST = ["DADA2_DENOISE_R1", "DADA2_DENOISE_R2", "DADA2_MERGE_PAIRS", "DADA2_10N_CONCAT"]
HASH_VALUE_COL_INDEX_LIST = [DADA2_DENOISE_R1_HASH_VALUE_COL_INDEX, DADA2_DENOISE_R2_HASH_VALUE_COL_INDEX,
                             DADA2_MERGE_PAIRS_HASH_VALUE_COL_INDEX, DADA2_10N_CONCAT_HASH_VALUE_COL_INDEX]


def get_csv_cells_content_map_by_column(file_path: str, start_row_index: int, key_col_index: int,
                                        value_col_index: int) -> dict:
    """
    Function to get all cell content in a specific column in a CSV file
    :param file_path: the file path of the CSV file
    :param start_row_index: the row index to start reading the file
    :param key_col_index: the column index of the key
    :param value_col_index: the column index of the value
    :return: a dictionary of the cell content
    e.g. {'trnLF_L5675_br01_F4121_br01': 'aad3b405edfeebe31da2f391fff81e5f'}
    """
    with open(file_path, newline='') as csvfile:
        reader = csv.reader(csvfile)
        try:
            # Skip to the desired row
            for _ in range(start_row_index):
                next(reader)

            # Read the row and get the content of the specific cell
            cell_content_map = dict()
            for row in reader:
                if row[value_col_index] == '-':
                    break
                sample_name = row[key_col_index]
                cell_content_map[sample_name] = row[value_col_index]
            return cell_content_map
        except IndexError:
            return None
        except StopIteration:
            return None


def get_batch_result_maps(batch_name: str, loci_name: str, start_row_index: int,
                          key_col_index: int, value_col_index: int,
                          total_run_number: int):
    """
    Function to get a list of result maps for a batch of runs
    """
    batch_result_map_list = []
    for i in range(total_run_number):
        file_path = f"{CSV_FILE_PATH_PREFIX}{batch_name}{i}{loci_name}{CSV_FILE_PATH_SUFFIX}"
        single_run_result = get_csv_cells_content_map_by_column(file_path, start_row_index, key_col_index,
                                                                value_col_index)
        batch_result_map_list.append(single_run_result)
    return batch_result_map_list


def __count_distinct_values(batch_result_map_list):
    """
    Function to count distinct values for each key in a list of result maps
    """
    aggregated_counts = {}
    for batch_result_map in batch_result_map_list:
        for key, value in batch_result_map.items():
            if key not in aggregated_counts:
                aggregated_counts[key] = set()
            aggregated_counts[key].add(value)
    return {key: len(values) for key, values in aggregated_counts.items()}


def __get_most_common_values(batch_result_map_list):
    """
    Function to extract the most common value for each key in a list of result maps
    """
    aggregated_map = {}
    for batch_result_map in batch_result_map_list:
        for key, value in batch_result_map.items():
            if key not in aggregated_map:
                aggregated_map[key] = Counter()
            aggregated_map[key].update([value])
    return {key: __hash_to_color(counter.most_common(1)[0][0]) if counter else None for key, counter in
            aggregated_map.items()}


def __hash_to_color(hash_str):
    """
    Function to convert a hash string to RGB color for visualization
    :param hash_str:
    :return: RGB color
    """
    if 'N/A' in hash_str:
        return 0, 0, 0

    # Convert the hash string to a number
    hash_num = int(hash_str, 16)

    # Extract RGB values from the hash number
    red = (hash_num & 0xFF0000) >> 16
    green = (hash_num & 0x00FF00) >> 8
    blue = hash_num & 0x0000FF

    return red, green, blue


def get_aggregated_csv_cells_content_counts_map_by_column_in_batch(batch_name: str, loci_name: str,
                                                                   start_row_index: int,
                                                                   key_col_index: int, value_col_index: int,
                                                                   total_run_number: int) -> dict:
    """
    Function to aggregate the cell content of a specific column into distinctive count in a batch of runs
    """
    batch_result_map_list = get_batch_result_maps(batch_name, loci_name, start_row_index,
                                                  key_col_index, value_col_index,
                                                  total_run_number)
    return __count_distinct_values(batch_result_map_list)


def get_aggregated_csv_cells_content_most_show_value_map_by_column_in_batch(batch_name: str, loci_name: str,
                                                                            start_row_index: int,
                                                                            key_col_index: int, value_col_index: int,
                                                                            total_run_number: int) -> dict:
    """
    Function to aggregate the cell content of a specific column into the highest frequency value in a batch of runs
    """
    batch_result_map_list = get_batch_result_maps(batch_name, loci_name, start_row_index,
                                                  key_col_index, value_col_index,
                                                  total_run_number)
    return __get_most_common_values(batch_result_map_list)


def __transpose_data_matrix(original_dict: dict) -> dict:
    """
    :param original_dict:
        {
        "SuperRed_35": {'trnLF_L5675_br01_F4121_br01': (248, 30, 95), 'trnLF_L5675_br01_F4121_br02': (223, 149, 71)},
        "filtered": {'trnLF_L5675_br01_F4121_br01': (248, 30, 95), 'trnLF_L5675_br01_F4121_br02': (13, 146, 228)},
        "filtered_selected_pure": {'trnLF_L5675_br01_F4121_br01': (248, 30, 95), 'trnLF_L5675_br01_F4121_br02': (223, 149, 71)},
        "no_learn_error": {'trnLF_L5675_br01_F4121_br01': (248, 30, 95), 'trnLF_L5675_br01_F4121_br02': (223, 149, 71)}
        }
    :return:
        {
        "trnLF_L5675_br01_F4121_br01": {"SuperRed_35": (248, 30, 95), "filtered": (248, 30, 95), "filtered_selected_pure": (248, 30, 95), "no_learn_error": (248, 30, 95)},
        "trnLF_L5675_br01_F4121_br02": {"SuperRed_35": (223, 149, 71), "filtered": (13, 146, 228), "filtered_selected_pure": (223, 149, 71), "no_learn_error": (223, 149, 71)}
        }
    """
    transposed_dict = {}
    for outer_key, inner_dict in original_dict.items():
        for inner_key, inner_value in inner_dict.items():
            if inner_key not in transposed_dict:
                transposed_dict[inner_key] = {}
            transposed_dict[inner_key][outer_key] = inner_value
    return transposed_dict


def csv_to_excel_with_background_color(csv_file: str, excel_file: str):
    """
    Convert a CSV file to an Excel file with background color
    e.g. (0, 0, 0) is the color for white
    :param csv_file: the file path of the CSV file
    :param excel_file: the file path of the Excel file
    :return: None
    """
    # Load CSV file and create a workbook
    with open(csv_file, 'r') as file:
        reader = csv.reader(file)
        wb = Workbook()
        ws = wb.active

        # Iterate over each row in the CSV file
        for row_index, row in enumerate(reader, start=1):
            # Iterate over each cell in the row
            for col_index, cell_value in enumerate(row, start=1):
                # Parse RGB value from the cell
                try:
                    rgb_values = tuple(map(int, cell_value.strip('()').split(',')))
                except ValueError:
                    # If parsing fails, fill the cell with string
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.value = cell_value
                    continue

                # Set background color and value of the cell
                cell = ws.cell(row=row_index, column=col_index)
                cell.fill = PatternFill(start_color=f'FF{rgb_values[0]:02X}{rgb_values[1]:02X}{rgb_values[2]:02X}',
                                        end_color=f'FF{rgb_values[0]:02X}{rgb_values[1]:02X}{rgb_values[2]:02X}',
                                        fill_type='solid')
                cell.value = ''  # add cell_value to show the color if needed
    # Save the Excel file
    wb.save(excel_file)


if __name__ == '__main__':
    csv_file_batch_run_name_list = [
        'SuperRed_35',
        'filtered',
        'filtered_selected_pure',
        'no_learn_error'
    ]

    csv_file_path_midfield = '/trnLF'
    batch_run_number = 15

    result = get_csv_cells_content_map_by_column(
        CSV_FILE_PATH_PREFIX + csv_file_batch_run_name_list[0] + '0' + csv_file_path_midfield + CSV_FILE_PATH_SUFFIX,
        FIRST_ASV_ROW_INDEX, ASV_BARCODE_COL_INDEX, DADA2_DENOISE_R1_HASH_VALUE_COL_INDEX)
    # print(result)

    result_matrix = dict()
    for index_number in range(len(HASH_VALUE_COL_INDEX_LIST)):
        for csv_file_batch_run_name in csv_file_batch_run_name_list:
            concat_key = HASH_VALUE_COL_NAME_LIST[index_number] + '_' + csv_file_batch_run_name
            value = get_aggregated_csv_cells_content_most_show_value_map_by_column_in_batch(csv_file_batch_run_name,
                                                                                            csv_file_path_midfield,
                                                                                            FIRST_ASV_ROW_INDEX,
                                                                                            ASV_BARCODE_COL_INDEX,
                                                                                            HASH_VALUE_COL_INDEX_LIST[
                                                                                                index_number],
                                                                                            batch_run_number)
            result_matrix[concat_key] = value
            # print(f"{concat_key}")
            # print(f"{result_matrix[concat_key]}")

    transposed_result_matrix = __transpose_data_matrix(result_matrix)
    print(transposed_result_matrix)

    csv_file_batch_run_name_and_col_name_list = []
    for HASH_VALUE_COL_NAME in HASH_VALUE_COL_NAME_LIST:
        for csv_file_batch_run_name in csv_file_batch_run_name_list:
            csv_file_batch_run_name_and_col_name_list.append(HASH_VALUE_COL_NAME + '_' + csv_file_batch_run_name)

    with open(CONCAT_CSV_FILE_PATH, mode='w', newline='') as target_csvfile:
        writer = csv.writer(target_csvfile)

        # Write the header row
        first_row = []
        second_row = []
        for hash_value_col_name in HASH_VALUE_COL_NAME_LIST:
            first_row.append(hash_value_col_name)
            first_row.append('')
            first_row.append('')
            first_row.append('')
            for csv_file_batch_run_name in csv_file_batch_run_name_list:
                second_row.append(csv_file_batch_run_name)
        writer.writerow([''] + first_row)
        writer.writerow([''] + second_row)

        # filled csv_file_batch_run_name_and_col_name_list with data by barcode name
        for barcode_name, result_map_by_barcode_name in transposed_result_matrix.items():
            """
            order by csv_file_batch_run_name_and_col_name_list, if the key is not in the result_map, then fill in '-'
            """
            target_row = []
            for csv_file_batch_run_name_and_col_name in csv_file_batch_run_name_and_col_name_list:
                if csv_file_batch_run_name_and_col_name in result_map_by_barcode_name.keys():
                    target_row.append(result_map_by_barcode_name[csv_file_batch_run_name_and_col_name])
                else:
                    target_row.append('-')
            writer.writerow([barcode_name] + target_row)

    # Convert the CSV file to an Excel file with background color
    csv_to_excel_with_background_color(CONCAT_CSV_FILE_PATH, CONCAT_EXCEL_FILE_PATH)

    print("CSV file 'matrix_data.csv' has been created successfully.")
